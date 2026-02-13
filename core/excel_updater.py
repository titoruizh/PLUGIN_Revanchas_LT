# -*- coding: utf-8 -*-
import os
import shutil
import datetime
import openpyxl
from openpyxl.chart import Reference

class ExcelUpdater:
    """
    Handles updating of the Excel template with new measurement data.
    Preserves existing formulas and formatting.
    """

    # Configuration for Muro Principal (MP)
    # PK at Column A (1-based index = 1)
    # Rows 13 to 85
    WALL_CONFIG = {
        "Muro Principal": {
            "sheet_name": "REV_MP", 
            "rows_range": (13, 200), # Extended range to catch all rows
            "pk_col_idx": 9, # Column I
            
            "shift_map": [
                ("C", "Q"), # Coronamiento -> Prev Crow
                ("F", "O"), # Lama -> Prev Lama
                ("E", "M"), # Revancha -> Prev Rev
            ],
            
            "write_map": {
                "crown": "C", 
                "lama": "F", 
                "width": "H", 
            },

            # Columns for calculated values
            # ⚠️ IMPORTANT NOTE ON FORMULAS VS STATIC VALUES:
            # We are intentionally calculating these values in Python and writing them as STATIC NUMBERS
            # instead of writing Excel formulas (e.g. '=C13-F13').
            # 
            # REASON:
            # When openpyxl writes a formula, it does NOT cache the calculation result.
            # Downstream programmatic tools (like XLSX.js or other Python scripts) that read this Excel file
            # without opening it in Excel first will see the calculated value as None/Undefined.
            # To ensure interoperability, we pre-calculate in Python and write the literal number.
            # 
            # DOWNSIDE:
            # If a user manually changes "Coronamiento" in the Excel, "Revancha" will NOT update automatically
            # until the file is re-exported or the user manually re-enters the formula.
            "calc_map": {
                "revancha": "E",          # = Crown - Lama
                "geomembrana": "J",       # Read Only (Input)
                "prev_revancha": "M",     # Read Only (Shifted)
                "dist_geo_lama": "K",     # = Geomembrana - Lama
                "dist_geo_crown": "L",    # = Crown - Geomembrana
                "delta_rev": "N"          # = Revancha - Prev Revancha
            },
            
            "date_cell": "F6",
            "chart_name_contains": "Gráfico 1", 
            "chart_title_prefix": "Coronamiento vs Lama MP "
        },
        
        "Muro Oeste": {
            "sheet_name": "Hoja1",
            "rows_range": (10, 100), # Extended range
            "pk_col_idx": 9, 
            
            "shift_map": [
                ("B", "Q"), 
                ("F", "O"), 
                ("E", "M"), 
            ],
            
            "write_map": {
                "crown": "B", 
                "lama": "F",  
                "width": "H", 
            },

            "calc_map": {
                "revancha": "E",
                "geomembrana": "J",
                "prev_revancha": "M",
                "dist_geo_lama": "K",
                "dist_geo_crown": "L",
                "delta_rev": "N"
            },
            
            "date_cell": "F6",
            "chart_name_contains": "Gráfico 8", 
            "chart_title_prefix": "Coronamiento vs Lama MO "
        },

        "Muro Este": {
            "sheet_name": "Hoja1",
            "rows_range": (13, 100), # Extended range
            "pk_col_idx": 9, 
            
            "shift_map": [
                ("C", "Q"), 
                ("F", "O"), 
                ("E", "M"), 
            ],
            
            "write_map": {
                "crown": "C", 
                "lama": "F", 
                "width": "H", 
            },

            "calc_map": {
                "revancha": "E",
                "geomembrana": "J",
                "prev_revancha": "M",
                "dist_geo_lama": "K",
                "dist_geo_crown": "L",
                "delta_rev": "N"
            },
            
            "date_cell": "F6",
            "chart_name_contains": "Gráfico 1", 
            "chart_title_prefix": "Coronamiento vs Lama ME "
        }
    }

    def __init__(self, output_path, template_path=None):
        """
        Initialize the ExcelUpdater.
        :param output_path: Path where the final Excel file will be saved.
        :param template_path: Optional path to the source template. If provided, it will be copied to output_path.
        """
        self.output_path = output_path
        self.template_path = template_path

    def update_wall_data(self, wall_name, measurements_data, dem_filename):
        """
        Main method to update the Excel file.
        :param wall_name: Name of the wall (e.g. "Muro Principal")
        :param measurements_data: List of dicts with measurements [{'pk':..., 'crown':..., 'lama':..., 'width':...}]
        :param dem_filename: Filename of the DEM to extract date
        :return: (bool, message)
        """
        try:
            # 0. Copy Template if provided
            if self.template_path:
                if not os.path.exists(self.template_path):
                    return False, f"El archivo plantilla no existe: {self.template_path}"
                
                try:
                    shutil.copy2(self.template_path, self.output_path)
                    print(f"✅ Plantilla copiada a: {self.output_path}")
                except Exception as e:
                    return False, f"Error copiando plantilla: {e}"
            else:
                # If no template path, assume output_path already exists (legacy/overwrite mode)
                if not os.path.exists(self.output_path):
                     return False, f"El archivo Excel no existe: {self.output_path}"

            # Work on output_path
            excel_target = self.output_path

            # 1. Load Workbook (Formulas) - To Write
            wb = openpyxl.load_workbook(excel_target, data_only=False)
            
            # 1b. Load Workbook (Values) - To Read Shifting Data
            # We need this to get the calculated result of formulas (like Revancha)
            wb_values = openpyxl.load_workbook(excel_target, data_only=True)
            
            # 2. Get Configuration
            # Normalize wall name / Handle aliases
            target_key = None
            wall_lower = wall_name.lower()
            
            if "muro 1" in wall_lower or "principal" in wall_lower:
                target_key = "Muro Principal"
            elif "muro 2" in wall_lower or "oeste" in wall_lower or "mo" in wall_lower:
                target_key = "Muro Oeste"
            elif "muro 3" in wall_lower or "este" in wall_lower or "me" in wall_lower:
                target_key = "Muro Este"
                
            config = self.WALL_CONFIG.get(target_key)
            
            if not config:
                # Fallback to fuzzy match
                for key in self.WALL_CONFIG:
                    if key.lower() in wall_lower:
                        config = self.WALL_CONFIG[key]
                        break
            
            if not config:
                return False, f"No hay configuración definida para el muro: {wall_name} (Normalizado: {target_key})"

            # 3. Get Sheet
            # Try to find sheet by name
            sheet = None
            sheet_values = None
            
            # Explicit config check first
            if config["sheet_name"] in wb.sheetnames:
                 sheet = wb[config["sheet_name"]]
                 sheet_values = wb_values[config["sheet_name"]]
            else:
                # Fallback search strategy
                search_terms = ["rev_mp", "rev", "muro principal", "muro 1", "mp", "principal"]
                found_sheet_name = None
                
                # Check each sheet name against search terms
                # Prioritize EXACT matches first? No, case insensitive contains
                for term in search_terms:
                    for sname in wb.sheetnames:
                        if term in sname.lower():
                            found_sheet_name = sname
                            break
                    if found_sheet_name:
                        break
                
                if found_sheet_name:
                    sheet = wb[found_sheet_name]
                    sheet_values = wb_values[found_sheet_name]
                else:
                    # List available sheets for error message
                    available = ", ".join(wb.sheetnames)
                    return False, f"No se encontró hoja para '{wall_name}'. Busqué: {search_terms}. Hojas disponibles: [{available}]"

            print(f"   📂 Usando hoja: {sheet.title}")

            # 4. Extract Date from DEM Filename
            # Format: DEM_MP_250101 -> 01/01/2025
            date_str = self._extract_date_from_filename(dem_filename)
            print(f"   📅 Fecha extraída: {date_str}")

            # 5. Shift Data (Current -> Previous)
            # Read from sheet_values (calculated), Write to sheet (formulas preserved where needed)
            self._shift_data(sheet, sheet_values, config)

            # 6. Write New Data
            # Pass sheet_values to read PKs reliably
            self._write_new_data(sheet, sheet_values, config, measurements_data)

            # 7. Update Date Cell
            # Format: dd-mm-yyyy
            self._update_date_cell(sheet, config, date_str)

            # 8. Update Chart Title
            self._update_chart_title(sheet, config, date_str)
            
            # Close value workbook
            wb_values.close()

            # 9. Save
            # NO BACKUP CREATION - we are saving to a new file (Save As workflow)
            
            wb.save(excel_target)
            return True, f"Datos exportados exitosamente a:\n{os.path.basename(excel_target)}\n\nFecha: {date_str}"

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

    def _extract_date_from_filename(self, filename):
        """
        Extract date from format *_YYMMDD*
        Example: DEM_MP_250101 -> 01/01/2025
        """
        try:
            # Remove extension
            name = os.path.splitext(os.path.basename(filename))[0]
            parts = name.split('_')
            
            # Look for 6 digit part
            date_part = None
            for part in parts:
                if len(part) == 6 and part.isdigit():
                    date_part = part
                    break
            
            if date_part:
                yy = int(date_part[0:2])
                mm = int(date_part[2:4])
                dd = int(date_part[4:6])
                
                # Assume 20xx
                full_year = 2000 + yy
                return f"{dd:02d}-{mm:02d}-{full_year}"
            
            return datetime.datetime.now().strftime("%d-%m-%Y")
        except:
            return datetime.datetime.now().strftime("%d-%m-%Y")

    def _shift_data(self, sheet_target, sheet_source, config):
        """
        Move data from Current columns to Previous columns.
        Read from Source (Values), Write to Target (Static).
        
        NEW LOGIC: Iterate rows first, check PK validity, then shift.
        Stops after 5 consecutive invalid rows OR if footer detected.
        """
        start_row, end_row = config["rows_range"]
        pk_col_idx = config["pk_col_idx"]
        
        # Identify Crown column to check for footer text (e.g. "Revancha (m)")
        crown_col_letter = config["write_map"].get("crown", "C")
        crown_col_idx = self._col_to_int(crown_col_letter)
        
        empty_count = 0
        
        # Iterate Row FIRST to check for valid PK (skip footer/empty rows)
        for row in range(start_row, end_row + 1):
            
            # CHECK 1: Detect Footer Text in Crown Column
            if crown_col_idx:
                 crown_val = sheet_source.cell(row=row, column=crown_col_idx).value
                 if isinstance(crown_val, str):
                     t = crown_val.lower()
                     if "resumen" in t or "fecha" in t or "revancha" in t or "ancho (" in t or "parametro" in t:
                         print(f"   ℹ️ Deteniendo escaneo en fila {row} (Pie de página detectado: '{crown_val}')")
                         break
            
            # CHECK 2: Validate PK
            pk_val = sheet_source.cell(row=row, column=pk_col_idx).value
            
            # Reset count if we found a valid PK
            if self._normalize_pk(pk_val) is not None:
                empty_count = 0 
            else:
                empty_count += 1
                if empty_count >= 5:
                    print(f"   ℹ️ Deteniendo escaneo en fila {row} (5 filas sin PK válido consecutivas)")
                    break
                continue # Skip invalid row

            # If Valid Row, Shift Columns
            for src_col_output, tgt_col_output in config["shift_map"]:
                # Read from Source (Values)
                src_cell = f"{src_col_output}{row}"
                tgt_cell = f"{tgt_col_output}{row}"
                
                val = sheet_source[src_cell].value
                
                # Write to Target (Previous)
                # Ensure we write values, not formulas
                if val is not None:
                     sheet_target[tgt_cell].value = val

    def _write_new_data(self, sheet, sheet_values, config, measurements):
        """
        Write new measurements matching PK.
        Uses sheet_values to read PK reliably.
        """
        start_row, end_row = config["rows_range"]
        pk_col_idx = config["pk_col_idx"]
        
        # Identify Crown column for footer check
        crown_col_letter = config["write_map"].get("crown", "C")
        crown_col_idx = self._col_to_int(crown_col_letter)
        
        # Create a lookup map for measurements: PK -> Data
        # Use robust normalization
        meas_map = {}
        print("   🔍 Procesando mediciones de la App:")
        for m in measurements:
            pk_raw = m.get('PK')
            pk_val = self._normalize_pk(pk_raw)
            if pk_val is not None:
                meas_map[pk_val] = m
                print(f"      OK: '{pk_raw}' -> {pk_val:.2f}")
            else:
                print(f"      ERR: No se pudo normalizar PK '{pk_raw}'")

        print(f"   📊 Total mediciones válidas para exportar: {len(meas_map)}")

        # Iterate rows in Excel to find PKs and write data
        count_updated = 0
        for row in range(start_row, end_row + 1):
            
            # CHECK 1: Detect Footer Text in Crown Column
            # (Same logic as shift_data to be consistent)
            if crown_col_idx:
                 crown_val = sheet_values.cell(row=row, column=crown_col_idx).value
                 if isinstance(crown_val, str):
                     t = crown_val.lower()
                     if "resumen" in t or "fecha" in t or "revancha" in t or "ancho (" in t or "parametro" in t:
                         print(f"   ℹ️ Deteniendo escritura en fila {row} (Pie de página detectado)")
                         break
            
            # Read PK from Excel (VALUES sheet to handle formulas/formatting)
            # Column I is index 9
            pk_cell = sheet_values.cell(row=row, column=pk_col_idx)
            pk_raw = pk_cell.value
            
            pk_float = self._normalize_pk(pk_raw)
            
            if pk_float is not None:
                # Find closest PK in measurements
                found_data = None
                closest_dist = float('inf')
                
                for m_pk, m_data in meas_map.items():
                    dist = abs(m_pk - pk_float)
                    if dist < 0.5: # Increased tolerance from 0.15 to 0.5
                        if dist < closest_dist:
                            closest_dist = dist
                            found_data = m_data
                
                if found_data:
                    # Debug log for match
                    # print(f"      MATCH Row {row}: Excel PK {pk_float:.2f} ~ App PK {found_data.get('PK')} (Diff: {closest_dist:.3f})")
                    count_updated += 1

                    
                    # 1. Variables to hold values for calculation
                    val_crown = None
                    val_lama = None
                    val_geomembrana = None
                    val_prev_revancha = None
                    
                    # 2. Write Basic Data (Crown, Lama, Width) & Capture Values
                    
                    # Crown
                    if "crown" in config["write_map"] and found_data.get('Cota_Coronamiento') is not None:
                        col_letter = config["write_map"]["crown"]
                        val_crown = float(found_data['Cota_Coronamiento'])
                        sheet[f"{col_letter}{row}"].value = val_crown
                        
                    # Lama
                    if "lama" in config["write_map"] and found_data.get('Lama') is not None:
                        col_letter = config["write_map"]["lama"]
                        val_lama = float(found_data['Lama'])
                        sheet[f"{col_letter}{row}"].value = val_lama
                        
                    # Width
                    if "width" in config["write_map"] and found_data.get('Ancho') is not None:
                        col_letter = config["write_map"]["width"]
                        sheet[f"{col_letter}{row}"].value = float(found_data['Ancho'])
                    
                    # 3. Read Input Values for Calculation (Geomembrana, Prev Revancha)
                    if "calc_map" in config:
                        cmap = config["calc_map"]
                        
                        # Read Geomembrana (Static Input)
                        if "geomembrana" in cmap:
                            col_geo = cmap["geomembrana"]
                            # Read from sheet (or sheet_values, usually static so either works)
                            # Using sheet_values is safer if it was a formula, but user implies static input
                            cell_val = sheet.cell(row=row, column=self._col_to_int(col_geo)).value
                            if isinstance(cell_val, (int, float)):
                                val_geomembrana = float(cell_val)
                        
                        # Read Previous Revancha (Static, was just shifted or already there)
                        if "prev_revancha" in cmap:
                            col_prev = cmap["prev_revancha"]
                            cell_val = sheet.cell(row=row, column=self._col_to_int(col_prev)).value
                            if isinstance(cell_val, (int, float)):
                                val_prev_revancha = float(cell_val)

                        # 4. Calculate and Write Derived Values
                        
                        # Revancha = Crown - Lama
                        if "revancha" in cmap and val_crown is not None and val_lama is not None:
                            val_rev = val_crown - val_lama
                            sheet[f"{cmap['revancha']}{row}"].value = val_rev
                            
                            # Delta Rev = Revancha - Prev Revancha
                            if "delta_rev" in cmap and val_prev_revancha is not None:
                                val_delta = val_rev - val_prev_revancha
                                sheet[f"{cmap['delta_rev']}{row}"].value = val_delta

                        # Dist Geo vs Lama = Geomembrana - Lama
                        if "dist_geo_lama" in cmap and val_geomembrana is not None and val_lama is not None:
                            val_dist_gl = val_geomembrana - val_lama
                            sheet[f"{cmap['dist_geo_lama']}{row}"].value = val_dist_gl
                            
                        # Dist Geo vs Crown = Crown - Geomembrana
                        if "dist_geo_crown" in cmap and val_crown is not None and val_geomembrana is not None:
                            val_dist_gc = val_crown - val_geomembrana
                            sheet[f"{cmap['dist_geo_crown']}{row}"].value = val_dist_gc

                    print(f"      MATCH: Row {row} (PK {pk_float:.2f}) updated")

        print(f"   ✅ Filas actualizadas en Excel: {count_updated}")

    def _col_to_int(self, col_str):
        """Convert column letter to index (A->1, B->2)"""
        if not col_str: return 0
        num = 0
        for c in col_str:
            if c in str(openpyxl.utils.cell.column_index_from_string(c)):
                 # already handle by library or manually
                 pass
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def _normalize_pk(self, value):
        """
        Convert various PK formats to float distance (meters).
        - 120.5 -> 120.5
        - "120.5" -> 120.5
        - "0+120.5" -> 120.5
        - "1+200" -> 1200.0
        """
        if value is None:
            return None
            
        try:
            # If already float/int
            if isinstance(value, (int, float)):
                return float(value)
            
            s = str(value).strip().replace(',', '.')
            
            if '+' in s:
                parts = s.split('+')
                if len(parts) == 2:
                    km = float(parts[0])
                    m = float(parts[1])
                    return km * 1000.0 + m
            
            return float(s)
        except:
            return None

    def _update_date_cell(self, sheet, config, date_str):
        if "date_cell" in config:
            sheet[config["date_cell"]].value = date_str

    def _update_chart_title(self, sheet, config, date_str):
        if "chart_name_contains" not in config:
            return
            
        print(f"   📊 Intentando actualizar título de gráfico...")
        try:
             # Iterate all charts
             count = 0
             for chart in sheet._charts:
                 # Try to force set title
                 new_title = config["chart_title_prefix"] + date_str
                 
                 # Debug info
                 try:
                     old_title = "Unknown"
                     if chart.title:
                         if hasattr(chart.title, "tx") and hasattr(chart.title.tx, "rich"):
                             old_title = "RichText"
                         elif hasattr(chart.title, "text"):
                             old_title = chart.title.text
                         else:
                             old_title = str(chart.title)
                     print(f"      Chart found. Old Title: {old_title}")
                 except:
                     pass

                 try:
                     # 1. Direct assignment (Best for simple titles)
                     chart.title = new_title
                     print(f"      ✅ Título actualizado a: {new_title}")
                     count += 1
                 except Exception as e:
                     print(f"      ⚠️ Falló asignación directa: {e}")
                     # 2. Try to navigate object if it exists?
                     # Openpyxl handling of titles is inconsistent across versions/chart types
                     pass
             
             if count == 0:
                 print("   ⚠️ No se encontraron gráficos iterables en sheet._charts")
                 
        except Exception as e:
            print(f"   ⚠️ Error general actualizando gráfico: {e}")
